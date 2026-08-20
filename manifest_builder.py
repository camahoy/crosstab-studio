"""
manifest_builder.py — Custom Cuts for Crosstab Studio
Takes a cuts spec CSV + current-wave banner(s) + optional previous-wave banner(s),
extracts specified metrics per slide, writes structured Excel output.
"""

import io, re, math
import csv
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font as XLFont, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import engine as _engine
from profiles import get_profile, get_profile_names


# ── Banner pre-parser — opens each file ONCE ──────────────────────────────────

class BannerCache:
    """
    Pre-parses all sheets from a banner file in a single pass.
    Holds parsed data in memory; no repeated file opens.
    """
    def __init__(self, file_bytes, profile_name, col_indices,
                 prefixes_needed=None, progress_cb=None):
        self.profile_name = profile_name
        self.col_indices  = col_indices

        # Read TOC / question index
        toc = _engine.read_toc(file_bytes, profile_name)
        if toc:
            self._toc = toc
            prefix_to_entries = {}
            for e in toc:
                prefix_to_entries.setdefault(e["prefix"], []).append(e)
        else:
            groups = _engine.fast_scan(file_bytes, profile_name)
            self._toc = []
            prefix_to_entries = {}
            for g in groups:
                for i, si in enumerate(g["sheets"]):
                    t = g.get("types", ["standard"] * len(g["sheets"]))[i]
                    entry = {"sheet_idx": si, "prefix": g["prefix"],
                             "wording": g["wording"], "sheet_type": t}
                    self._toc.append(entry)
                    prefix_to_entries.setdefault(g["prefix"], []).append(entry)

        self._prefix_map = prefix_to_entries  # prefix -> [entry]

        # Only parse sheets we actually need (those referenced by the manifest)
        if prefixes_needed is not None:
            needed = [e for e in self._toc if e["prefix"] in prefixes_needed]
        else:
            needed = self._toc

        all_indices = sorted({e["sheet_idx"] for e in needed})
        self._parsed = {}  # sheet_idx -> parsed dict

        # Open the Excel file ONCE and reuse it for every sheet parse
        xl = pd.ExcelFile(io.BytesIO(file_bytes))
        for i, si in enumerate(all_indices):
            try:
                p = _engine.parse_sheet_from_xl(xl, si, profile_name, col_indices)
                if p:
                    self._parsed[si] = p
            except Exception:
                pass
            if progress_cb:
                progress_cb(i + 1, len(all_indices))

    def prefixes(self):
        return set(self._prefix_map.keys())

    def get_columns(self):
        return self.col_indices

    def entries_for(self, prefix):
        """[(sheet_idx, wording, sheet_type)] for a prefix."""
        return [(e["sheet_idx"], e["wording"], e["sheet_type"])
                for e in self._prefix_map.get(prefix, [])]

    def parsed_for(self, prefix):
        """List of parsed dicts for a prefix (non-empty answers only)."""
        results = []
        for e in self._prefix_map.get(prefix, []):
            p = self._parsed.get(e["sheet_idx"])
            if p and p.get("answers"):
                results.append(p)
        return results

    def parsed_sheet(self, sheet_idx):
        return self._parsed.get(sheet_idx)

# ── Styles ────────────────────────────────────────────────────────────────────

def _f(bold=False, size=10, color="000000", name="Arial"):
    return XLFont(bold=bold, size=size, color=color, name=name)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(color="D0D7DE"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

SECTION_FILL  = _fill("0F2D4A")
SECTION_FONT  = _f(bold=True, size=12, color="FFFFFF")
SLIDE_FILL    = _fill("1A6EBD")
SLIDE_FONT    = _f(bold=True, size=10, color="FFFFFF")
HEADER_FILL   = _fill("E8F0FE")
HEADER_FONT   = _f(bold=True, size=9, color="0F2D4A")
W3_FILL       = _fill("F0F9FF")
W3_FONT       = _f(size=9, color="374151")
DELTA_FILL    = _fill("FFFBEB")
DELTA_FONT    = _f(bold=True, size=9, color="7C2D12")
BODY_FONT     = _f(size=9)
LABEL_FONT    = _f(bold=True, size=9)
MISS_FILL     = _fill("FEF2F2")
NEW_FILL      = _fill("F0FDF4")
BORDER        = _border()
THIN_BORDER   = _border("E2E8F0")

CTR = _align("center")
LEFT = _align("left")


# ── CSV parser ────────────────────────────────────────────────────────────────

def parse_manifest_csv(csv_bytes):
    """Parse manifest CSV bytes → list of row dicts."""
    text = csv_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for r in reader:
        slide = r.get("Slide", "").strip()
        if not slide:
            continue
        layout_raw = r.get("Table_Layout", "").strip().lower()
        rows.append({
            "slide":        slide,
            "section":      r.get("Section", "").strip(),
            "title":        r.get("Slide_Title", "").strip(),
            "prefix":       r.get("Question_Code", "").strip(),
            "sheet_type":   r.get("Sheet_Type", "").strip(),
            "metric":       r.get("Metric", "").strip(),
            "brands":       [b.strip() for b in r.get("Brands_Entities", "").split("|") if b.strip() and b.strip().upper() != "N/A"],
            "audiences":    [a.strip() for a in r.get("Audiences", "").split("|") if a.strip()],
            "wave_compare": r.get("Wave_Compare", "").strip(),
            "notes":        r.get("Notes", "").strip(),
            # pivot = brands/entities as columns, statements as rows (matches most report tables)
            # standard = brands as rows, audiences as columns
            "table_layout": "pivot" if layout_raw == "pivot" else "standard",
        })
    return rows


# ── Audience / column matching ────────────────────────────────────────────────

_AUDIENCE_PATTERNS = {
    "Gen Pop":    ["gen pop", "general population", "total", "adult"],
    "Tech Elite": ["tech elite", "technology elite", "tech"],
    "AI Fan":     ["ai fan", "ai fans", "artificial intelligence fan"],
    "Gen Z":      ["gen z", "generation z", "18-29", "18–29", "genz"],
}

def _match_col_for_audience(cols, audience_name):
    """
    Find best column index (from get_columns result) for an audience label.
    cols = [(col_idx, name, sublabel), ...]
    Returns col_idx or None.
    """
    audience_lc = audience_name.lower()
    patterns = _AUDIENCE_PATTERNS.get(audience_name, [audience_lc])

    # First: exact match on col name
    for idx, name, sub in cols:
        if name.strip().lower() == audience_lc:
            return idx

    # Second: pattern substring match
    for pat in patterns:
        for idx, name, sub in cols:
            combined = (name + " " + sub).lower()
            if pat in combined:
                return idx

    # Third: fuzzy — any word from audience in column name
    words = audience_lc.split()
    for idx, name, sub in cols:
        col_lc = (name + " " + sub).lower()
        if any(w in col_lc for w in words if len(w) > 2):
            return idx

    return None


def get_audience_col_map(file_bytes, profile_name, audience_list):
    """
    Return {audience_name: col_idx} for each audience in audience_list.
    Missing audiences get None.
    """
    cols = _engine.get_columns(file_bytes, profile_name)
    return {aud: _match_col_for_audience(cols, aud) for aud in audience_list}


def get_total_col_map(file_bytes, profile_name, label):
    """
    Return {label: col_idx} where col_idx is the 'Total' column in this file.
    Used when each banner file represents one audience and all have a 'Total' column.
    """
    cols = _engine.get_columns(file_bytes, profile_name)
    # Find 'Total' column — try exact match then common patterns
    for idx, name, sub in cols:
        if name.strip().lower() == "total":
            return {label: idx}
    for idx, name, sub in cols:
        if "total" in (name + " " + sub).lower():
            return {label: idx}
    # Fallback to first column
    if cols:
        return {label: cols[0][0]}
    return {}


# ── TOC inventory ─────────────────────────────────────────────────────────────

def get_all_prefixes(file_bytes, profile_name):
    """Return set of all question prefixes found in the banner."""
    toc = _engine.read_toc(file_bytes, profile_name)
    if toc:
        return {e["prefix"] for e in toc}
    groups = _engine.fast_scan(file_bytes, profile_name)
    return {g["prefix"] for g in groups}


# ── Metric extraction ─────────────────────────────────────────────────────────

def _fmt(v):
    """Format a float value as percentage string."""
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    if isinstance(v, float):
        pct = round(v * 100)
        return pct
    if isinstance(v, str):
        return v
    return v


def _extract_metric(parsed, metric_type, brands_filter=None):
    """
    From a parsed sheet dict, extract the metric.
    Returns list of (row_label, value_list) where value_list aligns to col_indices.

    metric_type: "T2B", "T4B", "All options", "Full scale", "MEAN RANK",
                 "PCT correct", "Weighted rank score", "RANK", "T2B optimistic + T2B pessimistic"
    brands_filter: list of brand strings to keep (case-insensitive substring match), or None=all
    """
    answers = parsed.get("answers", [])
    values  = parsed.get("values", [])
    metric_lc = metric_type.lower()

    def _brand_match(label, brands):
        if not brands:
            return True
        ll = label.lower()
        return any(b.lower() in ll or ll in b.lower() for b in brands)

    # Full scale / All options → return all rows (filtered by brand if needed)
    if any(k in metric_lc for k in ["all options", "full scale"]):
        rows = []
        for i, ans in enumerate(answers):
            if _brand_match(ans, brands_filter):
                rows.append((ans, values[i] if i < len(values) else []))
        return rows

    # T2B → find row labeled "Top 2 Box" / "T2B"; fallback: sum top 2 non-net rows
    if "t2b" in metric_lc and "t4b" not in metric_lc:
        # Check for explicit T2B row
        for i, ans in enumerate(answers):
            al = ans.lower()
            if "top 2 box" in al or al.startswith("t2b") or al == "top 2":
                if brands_filter:
                    # The T2B row represents the whole question, not a brand
                    # In summary sheets, rows ARE brands → match brand
                    if not _brand_match(ans, brands_filter):
                        continue
                return [(ans, values[i] if i < len(values) else [])]

        # If no explicit T2B row, this might be a summary sheet where rows = brands
        # Return all rows that match brand filter
        rows = []
        for i, ans in enumerate(answers):
            if _brand_match(ans, brands_filter):
                rows.append((ans, values[i] if i < len(values) else []))
        if rows:
            return rows

        # Last resort: sum top 2 numeric rows
        numeric_rows = [(a, v) for a, v in zip(answers, values)
                        if v and any(isinstance(x, float) for x in v)]
        if len(numeric_rows) >= 2:
            a1, v1 = numeric_rows[0]
            a2, v2 = numeric_rows[1]
            n_c = max(len(v1), len(v2))
            combined = []
            for ci in range(n_c):
                x1 = v1[ci] if ci < len(v1) and isinstance(v1[ci], float) else 0.0
                x2 = v2[ci] if ci < len(v2) and isinstance(v2[ci], float) else 0.0
                combined.append(x1 + x2)
            return [("Top 2 Box (computed)", combined)]
        return []

    # T4B → find "Top 4 Box" row; fallback: sum top 4
    if "t4b" in metric_lc:
        for i, ans in enumerate(answers):
            al = ans.lower()
            if "top 4 box" in al or al.startswith("t4b") or al == "top 4":
                return [(ans, values[i] if i < len(values) else [])]
        numeric_rows = [(a, v) for a, v in zip(answers, values)
                        if v and any(isinstance(x, float) for x in v)]
        if len(numeric_rows) >= 4:
            combined = []
            n_c = max(len(r[1]) for r in numeric_rows[:4])
            for ci in range(n_c):
                total = sum(
                    r[1][ci] if ci < len(r[1]) and isinstance(r[1][ci], float) else 0.0
                    for r in numeric_rows[:4]
                )
                combined.append(total)
            return [("Top 4 Box (computed)", combined)]
        return []

    # Mean / rank / weighted rank score
    if any(k in metric_lc for k in ["mean", "rank", "pct correct", "weighted rank"]):
        for i, ans in enumerate(answers):
            al = ans.lower()
            if any(k in al for k in ["mean", "average", "weighted rank", "rank score"]):
                return [(ans, values[i] if i < len(values) else [])]
        # Fallback: return all rows filtered by brand
        rows = []
        for i, ans in enumerate(answers):
            if _brand_match(ans, brands_filter):
                rows.append((ans, values[i] if i < len(values) else []))
        return rows

    # Default: return all rows filtered by brand
    rows = []
    for i, ans in enumerate(answers):
        if _brand_match(ans, brands_filter):
            rows.append((ans, values[i] if i < len(values) else []))
    return rows


def _find_brand_in_wording(wording, brands):
    """Return which brand from the list appears in this sheet's wording, or None."""
    wl = wording.lower()
    for b in brands:
        if b.lower() in wl:
            return b
    return None


# ── Slide data extraction ─────────────────────────────────────────────────────

def extract_slide_data(manifest_row, cache, col_map):
    """
    Extract data for one manifest row using a pre-built BannerCache.
    col_map = {audience: col_idx}

    Returns:
        {"found": bool, "rows": [(label, {aud: val}), ...], "base": {aud: n}, "wording": str}
    """
    prefix     = manifest_row["prefix"]
    metric     = manifest_row["metric"]
    brands     = manifest_row["brands"]
    audiences  = manifest_row["audiences"]
    sheet_type = manifest_row["sheet_type"]

    # Map audiences to column positions within the cache's col_indices
    cache_cols  = cache.get_columns()
    aud_order   = []
    col_pos     = []   # position index into cache_cols for each audience
    for aud in audiences:
        ci = col_map.get(aud)
        if ci is not None and ci in cache_cols:
            aud_order.append(aud)
            col_pos.append(cache_cols.index(ci))

    if not aud_order:
        return {"found": False, "rows": [], "base": {}, "wording": ""}

    per_entity = any(k in sheet_type.lower() for k in [
        "per brand", "per ceo", "per tool", "per assistant",
        "per statement", "per medium", "per outlet"])

    def _vals_for_row(full_vals):
        """Slice a full value row to just the requested audience positions."""
        return [full_vals[p] if p < len(full_vals) else None for p in col_pos]

    if per_entity and brands:
        all_sheets = cache.entries_for(prefix)
        result_rows  = []
        base_vals    = {}
        wording_seen = ""

        for brand in brands:
            matched_si = None
            for si, word, stype in all_sheets:
                if _find_brand_in_wording(word, [brand]):
                    matched_si   = si
                    wording_seen = word
                    break

            if matched_si is None:
                result_rows.append((brand, {aud: None for aud in aud_order}))
                continue

            p = cache.parsed_sheet(matched_si)
            if not p:
                result_rows.append((brand, {aud: None for aud in aud_order}))
                continue

            if not wording_seen:
                wording_seen = p.get("wording", "")

            for ai, aud in enumerate(aud_order):
                if aud not in base_vals:
                    bv = p["base_vals"][col_pos[ai]] if col_pos[ai] < len(p["base_vals"]) else None
                    if bv is not None:
                        base_vals[aud] = int(bv) if isinstance(bv, float) else bv

            # Build a view of parsed with only the requested columns
            p_view = dict(p)
            p_view["values"]    = [_vals_for_row(v) for v in p.get("values", [])]
            p_view["base_vals"] = _vals_for_row(p.get("base_vals", []))

            rows = _extract_metric(p_view, metric, brands_filter=None)
            if rows:
                _, vals = rows[0]
                row_dict = {aud: _fmt(vals[ai]) for ai, aud in enumerate(aud_order) if ai < len(vals)}
                result_rows.append((brand, row_dict))
            else:
                result_rows.append((brand, {aud: None for aud in aud_order}))

        return {"found": bool(result_rows), "rows": result_rows,
                "base": base_vals, "wording": wording_seen}

    else:
        parsed_list = cache.parsed_for(prefix)
        if not parsed_list:
            return {"found": False, "rows": [], "base": {}, "wording": ""}

        p = parsed_list[0]
        wording = p.get("wording", "")

        base_vals = {}
        for ai, aud in enumerate(aud_order):
            bv = p["base_vals"][col_pos[ai]] if col_pos[ai] < len(p.get("base_vals", [])) else None
            if bv is not None:
                base_vals[aud] = int(bv) if isinstance(bv, float) else bv

        p_view = dict(p)
        p_view["values"]    = [_vals_for_row(v) for v in p.get("values", [])]
        p_view["base_vals"] = _vals_for_row(p.get("base_vals", []))

        rows = _extract_metric(p_view, metric, brands_filter=brands or None)
        result_rows = []
        for label, vals in rows:
            row_dict = {aud: _fmt(vals[ai]) for ai, aud in enumerate(aud_order) if ai < len(vals)}
            result_rows.append((label, row_dict))

        return {"found": True, "rows": result_rows, "base": base_vals, "wording": wording}


# ── Grid slide extraction (brands/CEOs as columns, statements as rows) ─────────

def extract_grid_slide(manifest_row, w4_caches, w4_col_maps,
                       w3_caches=None, w3_col_maps=None):
    """
    For Grid_ prefixes: each entity has its own sheet with statements as rows.
    Returns data structured for a pivot table: entities (brands/CEOs) are columns,
    statement rows are rows.

    Returns dict per audience:
    {
      audience: {
        "statements": [str],
        "brands": [str],
        "base": {brand: n},
        "w4": {statement: {brand: val}},
        "w3": {statement: {brand: val}} or {},
      }
    }
    """
    prefix    = manifest_row["prefix"]
    brands    = manifest_row["brands"]
    audiences = manifest_row["audiences"]
    metric    = manifest_row["metric"]

    def _extract_for_cache(caches, col_maps):
        # {brand: {statement: {aud: val}}, base: {brand: {aud: n}}}
        brand_stmts  = {b: {} for b in brands}
        brand_base   = {b: {} for b in brands}
        stmt_order   = []

        for fi, cache in enumerate(caches):
            cm        = col_maps[fi] if fi < len(col_maps) else {}
            cache_cols = cache.get_columns()

            # aud → position in parsed values list
            aud_pos = {}
            for aud in audiences:
                ci = cm.get(aud)
                if ci is not None and ci in cache_cols:
                    aud_pos[aud] = cache_cols.index(ci)

            if not aud_pos:
                continue

            all_sheets = cache.entries_for(prefix)

            for brand in brands:
                matched_si = None
                for si, word, stype in all_sheets:
                    if _find_brand_in_wording(word, [brand]):
                        matched_si = si
                        break

                if matched_si is None:
                    continue

                p = cache.parsed_sheet(matched_si)
                if not p:
                    continue

                answers = p.get("answers", [])
                values  = p.get("values", [])
                bv_row  = p.get("base_vals", [])

                for aud, pos in aud_pos.items():
                    bv = bv_row[pos] if pos < len(bv_row) else None
                    if bv is not None and aud not in brand_base[brand]:
                        brand_base[brand][aud] = int(bv) if isinstance(bv, float) else bv

                for i, stmt in enumerate(answers):
                    if stmt not in stmt_order:
                        stmt_order.append(stmt)
                    if stmt not in brand_stmts[brand]:
                        brand_stmts[brand][stmt] = {}
                    vals = values[i] if i < len(values) else []
                    for aud, pos in aud_pos.items():
                        if aud not in brand_stmts[brand][stmt]:
                            v = vals[pos] if pos < len(vals) else None
                            brand_stmts[brand][stmt][aud] = _fmt(v)

        return stmt_order, brand_stmts, brand_base

    w4_stmts, w4_brand, w4_base = _extract_for_cache(w4_caches, w4_col_maps)
    w3_stmts, w3_brand, w3_base = (
        _extract_for_cache(w3_caches or [], w3_col_maps or [])
        if w3_caches else ([], {b: {} for b in brands}, {b: {} for b in brands})
    )

    # Organise per audience
    result = {}
    for aud in audiences:
        w4_tbl = {}
        for stmt in w4_stmts:
            w4_tbl[stmt] = {b: w4_brand[b].get(stmt, {}).get(aud) for b in brands}
        w3_tbl = {}
        if w3_caches and w3_stmts:
            for stmt in w3_stmts:
                w3_tbl[stmt] = {b: w3_brand[b].get(stmt, {}).get(aud) for b in brands}
        result[aud] = {
            "statements": w4_stmts,
            "brands":     brands,
            "base":       {b: w4_base[b].get(aud) for b in brands},
            "w4":         w4_tbl,
            "w3":         w3_tbl,
        }
    return result


def _delta_str(v4, v3):
    """Format inline change indicator like W3 report style: '+3pp' / '-2pp' / '(-)' """
    if not isinstance(v4, int) or not isinstance(v3, int):
        return ""
    d = v4 - v3
    if d == 0:
        return " (-)"
    return f" (+{d}pp)" if d > 0 else f" ({d}pp)"


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_cuts_excel(manifest_rows, w4_files, w3_files, profile_name,
                     missing_prefixes, new_prefixes, progress_cb=None,
                     w4_labels=None, w3_labels=None):
    """
    Build the output Excel.
    w4_files / w3_files: list of {bytes, label, name}
    w4_labels / w3_labels: optional list of audience labels (one per file).
        When provided, each file's 'Total' column is mapped to its label.
        When None, audience columns are auto-detected by name matching.
    missing_prefixes: set of prefixes in manifest not found in W4
    new_prefixes: set of prefixes in W4 not in manifest
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Step 1: collect all unique audiences across all rows ──────────────────
    all_audiences = []
    seen_aud = set()
    for row in manifest_rows:
        for a in row["audiences"]:
            if a not in seen_aud:
                all_audiences.append(a)
                seen_aud.add(a)

    # ── Step 2: build col maps (audience → column index) per file ─────────────
    # If file labels provided, map each file's Total column to its label.
    # Otherwise auto-detect audience columns by name.
    if w4_labels and len(w4_labels) == len(w4_files):
        w4_col_maps = [get_total_col_map(f["bytes"], profile_name, w4_labels[i])
                       for i, f in enumerate(w4_files)]
    else:
        w4_col_maps = [get_audience_col_map(f["bytes"], profile_name, all_audiences)
                       for f in w4_files]

    if w3_files:
        if w3_labels and len(w3_labels) == len(w3_files):
            w3_col_maps = [get_total_col_map(f["bytes"], profile_name, w3_labels[i])
                           for i, f in enumerate(w3_files)]
        else:
            w3_col_maps = [get_audience_col_map(f["bytes"], profile_name, all_audiences)
                           for f in w3_files]
    else:
        w3_col_maps = []

    def _merged_col_map(col_maps, aud_list):
        result = {}
        for aud in aud_list:
            for cm in col_maps:
                if cm.get(aud) is not None:
                    result[aud] = cm[aud]
                    break
        return result

    # ── Step 3: pre-parse each banner file ONCE into a BannerCache ───────────
    # Build union of all col_indices needed across all files
    all_w4_cols = sorted({ci for cm in w4_col_maps for ci in cm.values() if ci is not None})
    all_w3_cols = sorted({ci for cm in w3_col_maps for ci in cm.values() if ci is not None})

    # Build one BannerCache per file — each opens the Excel once and parses all sheets
    n_w4 = len(w4_files)
    n_w3 = len(w3_files)
    total_files = n_w4 + n_w3

    def _sheet_progress(file_i, done, total):
        if progress_cb:
            # Scale sheet progress to 0→1 within one file's share
            file_frac = file_i / max(total_files, 1)
            next_frac = (file_i + 1) / max(total_files, 1)
            progress_cb(file_frac + (done / max(total, 1)) * (next_frac - file_frac), 1.0)

    spec_prefixes = {r["prefix"] for r in manifest_rows if r["prefix"]}

    w4_caches = []
    for fi, finfo in enumerate(w4_files):
        bc = BannerCache(finfo["bytes"], profile_name, all_w4_cols,
                         prefixes_needed=spec_prefixes,
                         progress_cb=lambda d, t, _fi=fi: _sheet_progress(_fi, d, t))
        w4_caches.append(bc)

    w3_caches = []
    for fi, finfo in enumerate(w3_files):
        bc = BannerCache(finfo["bytes"], profile_name, all_w3_cols,
                         prefixes_needed=spec_prefixes,
                         progress_cb=lambda d, t, _fi=fi: _sheet_progress(n_w4 + _fi, d, t))
        w3_caches.append(bc)

    # ── Step 4: group manifest rows by section and write Excel ────────────────
    sections = {}
    section_order = []
    for row in manifest_rows:
        sec = row["section"] or "Other"
        if sec not in sections:
            sections[sec] = []
            section_order.append(sec)
        sections[sec].append(row)

    total_slides = len(manifest_rows)
    slide_done   = 0
    for sec in section_order:
        tab_name = re.sub(r'[\\/*?\[\]:]', '', sec)[:31]
        ws = wb.create_sheet(title=tab_name)
        ws.column_dimensions["A"].width = 42
        current_row = 1

        # Section header
        for ci in range(1, 20):
            ws.cell(row=current_row, column=ci).fill = SECTION_FILL
        c = ws.cell(row=current_row, column=1, value=sec.upper())
        c.font = SECTION_FONT
        c.alignment = LEFT
        current_row += 2

        for mrow in sections[sec]:
            prefix     = mrow["prefix"]
            audiences  = mrow["audiences"]
            do_wave    = mrow["wave_compare"].lower().startswith("w") and bool(w3_files)
            is_missing = prefix in missing_prefixes
            is_new_q   = mrow["wave_compare"].lower() == "n/a - new"

            # Pivot layout: brands/CEOs as columns, statements as rows — matches report table format
            # Controlled explicitly by Table_Layout column in manifest CSV (value: "pivot")
            is_grid = mrow["table_layout"] == "pivot" and bool(mrow["brands"])

            # Merge data from ALL current-wave caches so subgroup banners contribute
            def _merge_cache_data(caches, col_maps):
                merged_rows = {}   # label → {aud: val}
                merged_base = {}   # aud → n
                wording = ""
                for fi, bc in enumerate(caches):
                    cm = col_maps[fi] if fi < len(col_maps) else {}
                    data = extract_slide_data(mrow, bc, cm)
                    if data["found"]:
                        wording = data["wording"] or wording
                        for label, val_dict in data["rows"]:
                            if label not in merged_rows:
                                merged_rows[label] = {}
                            merged_rows[label].update(val_dict)
                        merged_base.update(data["base"])
                return {
                    "found": bool(merged_rows),
                    "rows": list(merged_rows.items()),
                    "base": merged_base,
                    "wording": wording,
                }

            if not is_grid:
                w4_data = _merge_cache_data(w4_caches, w4_col_maps)
                w3_data = (
                    _merge_cache_data(w3_caches, w3_col_maps)
                    if do_wave and w3_caches
                    else {"found": False, "rows": [], "base": {}, "wording": ""}
                )

            # ── Slide header ──────────────────────────────────────────
            slide_text = f"Slide {mrow['slide']}: {mrow['title']}"
            if is_missing:
                slide_text += "  ⚠ NOT FOUND IN BANNER"
            elif is_new_q:
                slide_text += "  ★ NEW THIS WAVE"

            fill = MISS_FILL if is_missing else (NEW_FILL if is_new_q else None)
            c = ws.cell(row=current_row, column=1, value=slide_text)
            c.font = SLIDE_FONT if not is_missing else _f(bold=True, size=10, color="991B1B")
            c.fill = _fill("FEF2F2") if is_missing else (_fill("F0FDF4") if is_new_q else SLIDE_FILL)
            c.alignment = LEFT
            # Fill rest of header row
            n_header_cols = 1 + len(audiences) * (3 if do_wave else 1)
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=max(n_header_cols, 2))
            current_row += 1

            if mrow["notes"]:
                nc = ws.cell(row=current_row, column=1, value=mrow["notes"])
                nc.font = _f(size=8, color="6B7280")
                nc.alignment = LEFT
                current_row += 1

            if is_missing:
                current_row += 1
                continue

            # ── GRID layout: brands/CEOs as columns, statements as rows ───────
            if is_grid:
                grid = extract_grid_slide(
                    mrow, w4_caches, w4_col_maps,
                    w3_caches if do_wave else None,
                    w3_col_maps if do_wave else None,
                )

                for aud in audiences:
                    aud_grid = grid.get(aud, {})
                    stmts    = aud_grid.get("statements", [])
                    brands   = aud_grid.get("brands", mrow["brands"])
                    w4_tbl   = aud_grid.get("w4", {})
                    w3_tbl   = aud_grid.get("w3", {})
                    base_map = aud_grid.get("base", {})

                    # Audience sub-label
                    al = ws.cell(row=current_row, column=1, value=aud)
                    al.font   = _f(bold=True, size=9, color="0F2D4A")
                    al.fill   = HEADER_FILL
                    al.alignment = LEFT
                    current_row += 1

                    # Header row: blank | Brand1 | Brand2 | ...
                    ws.cell(row=current_row, column=1, value="").fill = HEADER_FILL
                    for bi, brand in enumerate(brands):
                        hc = ws.cell(row=current_row, column=2 + bi, value=brand)
                        hc.font  = HEADER_FONT
                        hc.fill  = HEADER_FILL
                        hc.alignment = CTR
                        ws.column_dimensions[get_column_letter(2 + bi)].width = max(
                            14, len(brand) + 2)
                    current_row += 1

                    # Base row
                    any_base = any(v is not None for v in base_map.values())
                    if any_base:
                        bc_cell = ws.cell(row=current_row, column=1, value="Base (n)")
                        bc_cell.font = _f(bold=True, size=8, color="6B7280")
                        bc_cell.alignment = LEFT
                        for bi, brand in enumerate(brands):
                            bv = base_map.get(brand)
                            if bv is not None:
                                c2 = ws.cell(row=current_row, column=2 + bi, value=bv)
                                c2.font = _f(size=8, color="6B7280")
                                c2.alignment = CTR
                        current_row += 1

                    # Statement rows
                    alt = False
                    for stmt in stmts:
                        row_fill = _fill("F6F8FA") if alt else _fill("FFFFFF")
                        alt = not alt

                        lc = ws.cell(row=current_row, column=1, value=stmt)
                        lc.font      = LABEL_FONT
                        lc.fill      = row_fill
                        lc.alignment = LEFT
                        lc.border    = THIN_BORDER

                        for bi, brand in enumerate(brands):
                            v4 = w4_tbl.get(stmt, {}).get(brand)
                            v3 = w3_tbl.get(stmt, {}).get(brand) if do_wave else None

                            delta = _delta_str(v4, v3) if do_wave else ""
                            disp  = (f"{v4}%{delta}" if isinstance(v4, int)
                                     else (str(v4) if v4 else "—"))

                            dc = ws.cell(row=current_row, column=2 + bi, value=disp)
                            dc.font      = BODY_FONT
                            dc.fill      = row_fill
                            dc.alignment = CTR
                            dc.border    = THIN_BORDER

                        current_row += 1

                    if not stmts:
                        nc = ws.cell(row=current_row, column=1, value="(no data extracted)")
                        nc.font = _f(size=9, color="9CA3AF")
                        current_row += 1

                    current_row += 1  # spacer between audiences

                if progress_cb:
                    progress_cb(slide_done, total_slides)
                slide_done += 1
                current_row += 1
                continue

            # ── Column headers ────────────────────────────────────────
            ws.cell(row=current_row, column=1, value="").fill = HEADER_FILL

            col_cursor = 2
            col_map_pos = {}  # audience → (w4_col, w3_col, delta_col) in worksheet

            for aud in audiences:
                if do_wave:
                    c_w4 = ws.cell(row=current_row, column=col_cursor, value=aud)
                    c_w4.font = HEADER_FONT
                    c_w4.fill = HEADER_FILL
                    c_w4.alignment = CTR
                    ws.column_dimensions[get_column_letter(col_cursor)].width = 12

                    c_w3 = ws.cell(row=current_row, column=col_cursor + 1, value=f"{aud} (Prev)")
                    c_w3.font = _f(bold=True, size=9, color="374151")
                    c_w3.fill = W3_FILL
                    c_w3.alignment = CTR
                    ws.column_dimensions[get_column_letter(col_cursor + 1)].width = 14

                    c_d = ws.cell(row=current_row, column=col_cursor + 2, value="Δ")
                    c_d.font = DELTA_FONT
                    c_d.fill = DELTA_FILL
                    c_d.alignment = CTR
                    ws.column_dimensions[get_column_letter(col_cursor + 2)].width = 7

                    col_map_pos[aud] = (col_cursor, col_cursor + 1, col_cursor + 2)
                    col_cursor += 3
                else:
                    c_w4 = ws.cell(row=current_row, column=col_cursor, value=aud)
                    c_w4.font = HEADER_FONT
                    c_w4.fill = HEADER_FILL
                    c_w4.alignment = CTR
                    ws.column_dimensions[get_column_letter(col_cursor)].width = 14

                    col_map_pos[aud] = (col_cursor, None, None)
                    col_cursor += 1

            current_row += 1

            # ── Base row ──────────────────────────────────────────────
            if w4_data["base"]:
                bc = ws.cell(row=current_row, column=1, value="Base (n)")
                bc.font = _f(bold=True, size=8, color="6B7280")
                bc.alignment = LEFT
                for aud in audiences:
                    pos = col_map_pos.get(aud)
                    if not pos:
                        continue
                    w4_col, w3_col, d_col = pos
                    bv4 = w4_data["base"].get(aud)
                    if bv4 is not None:
                        cell = ws.cell(row=current_row, column=w4_col, value=bv4)
                        cell.font = _f(size=8, color="6B7280")
                        cell.alignment = CTR
                    if do_wave and w3_col:
                        bv3 = w3_data["base"].get(aud)
                        if bv3 is not None:
                            cell = ws.cell(row=current_row, column=w3_col, value=bv3)
                            cell.font = _f(size=8, color="6B7280")
                            cell.alignment = CTR
                current_row += 1

            # ── Data rows ─────────────────────────────────────────────
            # Build lookup for W3 rows
            w3_lookup = {label: vals for label, vals in w3_data.get("rows", [])}

            alt = False
            for label, w4_vals in w4_data.get("rows", []):
                row_fill = _fill("F6F8FA") if alt else _fill("FFFFFF")
                alt = not alt

                lc = ws.cell(row=current_row, column=1, value=label)
                lc.font = LABEL_FONT
                lc.fill = row_fill
                lc.alignment = LEFT
                lc.border = THIN_BORDER

                for aud in audiences:
                    pos = col_map_pos.get(aud)
                    if not pos:
                        continue
                    w4_col, w3_col, d_col = pos

                    v4 = w4_vals.get(aud)
                    v3 = w3_lookup.get(label, {}).get(aud) if do_wave else None

                    # W4 value
                    cell4 = ws.cell(row=current_row, column=w4_col,
                                    value=f"{v4}%" if isinstance(v4, int) else (v4 or "—"))
                    cell4.font = BODY_FONT
                    cell4.fill = row_fill
                    cell4.alignment = CTR
                    cell4.border = THIN_BORDER

                    if do_wave and w3_col:
                        cell3 = ws.cell(row=current_row, column=w3_col,
                                        value=f"{v3}%" if isinstance(v3, int) else (v3 or "—"))
                        cell3.font = W3_FONT
                        cell3.fill = _fill("F0F9FF")
                        cell3.alignment = CTR
                        cell3.border = THIN_BORDER

                        if d_col and isinstance(v4, int) and isinstance(v3, int):
                            delta = v4 - v3
                            delta_str = f"+{delta}pp" if delta > 0 else (f"{delta}pp" if delta != 0 else "—")
                            dc = ws.cell(row=current_row, column=d_col, value=delta_str)
                            dc.font = _f(bold=True, size=9,
                                         color="16A34A" if delta > 0 else ("DC2626" if delta < 0 else "6B7280"))
                            dc.fill = DELTA_FILL
                            dc.alignment = CTR
                            dc.border = THIN_BORDER
                        elif d_col:
                            dc = ws.cell(row=current_row, column=d_col, value="—")
                            dc.font = _f(size=9, color="9CA3AF")
                            dc.fill = DELTA_FILL
                            dc.alignment = CTR

                current_row += 1

            if not w4_data.get("rows"):
                nc = ws.cell(row=current_row, column=1, value="(no data extracted)")
                nc.font = _f(size=9, color="9CA3AF")
                current_row += 1

            current_row += 1  # spacer between slides

            if progress_cb:
                progress_cb(slide_done, total_slides)
            slide_done += 1

    # ── Flags tab ─────────────────────────────────────────────────────────────
    ws_f = wb.create_sheet(title="⚑ Flags")
    ws_f.column_dimensions["A"].width = 20
    ws_f.column_dimensions["B"].width = 60

    r = 1
    hdr = ws_f.cell(row=r, column=1, value="TYPE")
    hdr.font = HEADER_FONT; hdr.fill = HEADER_FILL
    hdr2 = ws_f.cell(row=r, column=2, value="PREFIX / NOTE")
    hdr2.font = HEADER_FONT; hdr2.fill = HEADER_FILL
    r += 1

    for p in sorted(missing_prefixes):
        c1 = ws_f.cell(row=r, column=1, value="MISSING")
        c1.font = _f(bold=True, size=9, color="991B1B")
        c1.fill = MISS_FILL
        c2 = ws_f.cell(row=r, column=2, value=p)
        c2.font = _f(size=9)
        r += 1

    for p in sorted(new_prefixes):
        c1 = ws_f.cell(row=r, column=1, value="NEW THIS WAVE")
        c1.font = _f(bold=True, size=9, color="166534")
        c1.fill = NEW_FILL
        c2 = ws_f.cell(row=r, column=2, value=p)
        c2.font = _f(size=9)
        r += 1

    if not missing_prefixes and not new_prefixes:
        ws_f.cell(row=r, column=1, value="No flags — all prefixes matched.")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Streamlit UI ──────────────────────────────────────────────────────────────

def show_manifest_builder():
    st.markdown(
        "<div style='font-size:0.82rem;color:#374151;margin-bottom:16px'>"
        "Upload your cuts spec CSV and banner file(s) from the DP team. "
        "The tool extracts exactly the metrics you need and outputs clean, report-ready tables."
        "</div>",
        unsafe_allow_html=True,
    )

    col_l, col_r = st.columns([1, 1])

    with col_l:
        st.markdown('<div class="step-label">Cuts spec (CSV)</div>', unsafe_allow_html=True)
        st.markdown(
            "<div style='font-size:0.72rem;color:#6B7280;margin-bottom:6px'>"
            "CSV with columns: Slide, Section, Slide_Title, Question_Code, Sheet_Type, "
            "Metric, Brands_Entities, Audiences, Wave_Compare, Table_Layout, Notes — "
            "Table_Layout: <em>pivot</em> = brands/CEOs as columns, statements as rows; "
            "<em>standard</em> (default) = brands as rows, audiences as columns"
            "</div>", unsafe_allow_html=True,
        )
        manifest_file = st.file_uploader(
            "Upload cuts spec CSV",
            type=["csv"],
            key="mb_manifest",
            label_visibility="collapsed",
        )

    with col_r:
        st.markdown('<div class="step-label">Banner format</div>', unsafe_allow_html=True)
        profile_names = [p for p in get_profile_names() if p != "+ Add new format"]
        mb_profile = st.selectbox(
            "Profile",
            profile_names,
            key="mb_profile",
            label_visibility="collapsed",
            help="Select the banner format. Use 'Corporate Reputation' for GQR-style banners.",
        )

    st.markdown('<div class="step-label" style="margin-top:16px">Current wave banner(s)</div>', unsafe_allow_html=True)
    st.markdown(
        "<div style='font-size:0.72rem;color:#6B7280;margin-bottom:6px'>"
        "Upload one or more banner Excel files. If audiences come from separate files (e.g. a Total banner, a Gen Z banner, an AI Fans banner), "
        "upload all of them here and give each a label matching what you put in the <em>Audiences</em> column of your manifest."
        "</div>",
        unsafe_allow_html=True,
    )
    w4_uploads = st.file_uploader(
        "Current wave banners",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key="mb_w4",
        label_visibility="collapsed",
    )

    # Per-file audience labels (shown when >1 file or always for clarity)
    w4_file_labels = []
    if w4_uploads:
        if len(w4_uploads) > 1:
            st.markdown(
                "<div style='font-size:0.72rem;color:#374151;margin-top:6px;margin-bottom:4px'>"
                "<strong>Label each file</strong> — use the same audience name you put in the "
                "<em>Audiences</em> column of your manifest (e.g. <em>Gen Pop</em>, <em>Gen Z</em>). "
                "The tool maps each file's Total column to this label."
                "</div>",
                unsafe_allow_html=True,
            )
            label_cols = st.columns(min(len(w4_uploads), 4))
            for i, uf in enumerate(w4_uploads):
                with label_cols[i % 4]:
                    default = uf.name.replace(".xlsx", "").replace(".xls", "")[:20]
                    lbl = st.text_input(
                        uf.name[:25],
                        value=default,
                        key=f"mb_w4_label_{i}_{uf.name}",
                        help="Audience name this file represents — must match Audiences column in manifest.",
                    )
                    w4_file_labels.append(lbl.strip())
        else:
            # Single file — auto-detect columns normally
            w4_file_labels = []

    with st.expander("Previous wave banners — Trended (optional)", expanded=False):
        st.markdown(
            "<div style='font-size:0.72rem;color:#6B7280;margin-bottom:6px'>"
            "Upload previous-wave banner file(s) to add a Trended column and Δ change for each metric. "
            "Label each file to match its audience."
            "</div>",
            unsafe_allow_html=True,
        )
        w3_uploads = st.file_uploader(
            "Previous wave banners",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="mb_w3",
            label_visibility="collapsed",
        )
        w3_file_labels = []
        if w3_uploads and len(w3_uploads) > 1:
            st.markdown(
                "<div style='font-size:0.72rem;color:#374151;margin-top:6px;margin-bottom:4px'>"
                "Label each previous-wave file:</div>",
                unsafe_allow_html=True,
            )
            w3_label_cols = st.columns(min(len(w3_uploads), 4))
            for i, uf in enumerate(w3_uploads):
                with w3_label_cols[i % 4]:
                    default = uf.name.replace(".xlsx", "").replace(".xls", "")[:20]
                    lbl = st.text_input(
                        uf.name[:25],
                        value=default,
                        key=f"mb_w3_label_{i}_{uf.name}",
                    )
                    w3_file_labels.append(lbl.strip())

    if not manifest_file or not w4_uploads:
        st.info("Upload a cuts spec CSV and at least one current-wave banner to continue.")
        return

    manifest_rows = parse_manifest_csv(manifest_file.getvalue())
    if not manifest_rows:
        st.error("Could not parse the cuts spec CSV — check the file format.")
        return

    n_sections = len({r["section"] for r in manifest_rows})
    trended    = bool(w3_uploads)
    st.markdown(
        f'<span class="stat-pill">{len(manifest_rows)} cuts</span>'
        f'<span class="stat-pill">{n_sections} section{"s" if n_sections != 1 else ""}</span>'
        f'<span class="stat-pill">{len(w4_uploads)} banner file{"s" if len(w4_uploads) != 1 else ""}</span>'
        + (f'<span class="stat-pill">Trended ✓</span>' if trended else ""),
        unsafe_allow_html=True,
    )
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    with st.expander("Preview cuts spec", expanded=False):
        df_prev = pd.DataFrame([{
            "Slide":     r["slide"],
            "Section":   r["section"],
            "Title":     r["title"][:50],
            "Code":      r["prefix"],
            "Metric":    r["metric"],
            "Audiences": ", ".join(r["audiences"]),
            "Trended":   "Yes" if r["wave_compare"].lower().startswith("w") else "—",
        } for r in manifest_rows])
        st.dataframe(df_prev, use_container_width=True, hide_index=True)

    if st.button("◈  Build custom cuts Excel", key="mb_run"):
        w4_files = [{"bytes": f.getvalue(), "label": f.name, "name": f.name}
                    for f in w4_uploads]
        w3_files = [{"bytes": f.getvalue(), "label": f.name, "name": f.name}
                    for f in w3_uploads] if w3_uploads else []
        # Use per-file labels when multiple files were uploaded with explicit labels
        _w4_labels = w4_file_labels if len(w4_file_labels) == len(w4_files) and len(w4_files) > 1 else None
        _w3_labels = w3_file_labels if len(w3_file_labels) == len(w3_files) and len(w3_files) > 1 else None

        prog_area = st.empty()
        status_area = st.empty()
        try:
            status_area.markdown(
                "<div style='font-size:0.82rem;color:#374151'>Step 1 of 3 — Reading banner TOC…</div>",
                unsafe_allow_html=True,
            )
            prog_area.progress(0.05)

            all_w4_prefixes = set()
            for finfo in w4_files:
                all_w4_prefixes |= get_all_prefixes(finfo["bytes"], mb_profile)

            prog_area.progress(0.2)
            status_area.markdown(
                "<div style='font-size:0.82rem;color:#374151'>Step 2 of 3 — Matching columns…</div>",
                unsafe_allow_html=True,
            )

            spec_prefixes    = {r["prefix"] for r in manifest_rows if r["prefix"]}
            missing_prefixes = spec_prefixes - all_w4_prefixes
            new_prefixes     = all_w4_prefixes - spec_prefixes

            if missing_prefixes:
                st.warning(
                    f"⚠ {len(missing_prefixes)} code(s) in your spec not found in banner: "
                    + ", ".join(sorted(missing_prefixes))
                )
            if new_prefixes:
                st.info(
                    f"★ {len(new_prefixes)} code(s) in the banner not in your spec (new this wave): "
                    + ", ".join(sorted(new_prefixes))
                )

            prog_area.progress(0.3)
            status_area.markdown(
                f"<div style='font-size:0.82rem;color:#374151'>Step 3 of 3 — Parsing {len(w4_files) + len(w3_files)} banner file(s)…</div>",
                unsafe_allow_html=True,
            )

            def _build_progress(frac, _total):
                prog_area.progress(min(0.3 + 0.65 * frac, 0.95))

            result_bytes = build_cuts_excel(
                manifest_rows, w4_files, w3_files, mb_profile,
                missing_prefixes, new_prefixes,
                progress_cb=_build_progress,
                w4_labels=_w4_labels,
                w3_labels=_w3_labels,
            )

            prog_area.progress(1.0)
            status_area.empty()
            prog_area.empty()

            st.success(f"Done — {len(manifest_rows)} cuts across {n_sections} sections")
            st.download_button(
                label="⬇  Download custom cuts Excel",
                data=result_bytes,
                file_name="custom_cuts.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="mb_download",
            )

        except Exception as e:
            import traceback
            prog_area.empty()
            status_area.empty()
            st.error(f"Build failed: {e}")
            st.code(traceback.format_exc())
